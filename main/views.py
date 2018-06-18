from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import auth
from main.models import Exam, Answer, Question, Test, Subdivision, Worker
from django.contrib.auth.models import User
from testing import choosing_questions
import datetime
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

# Create your views here.

def index(request):
    context = {}

    return render(request, 'main/index.html', context)

def login(request):
    if request.POST:
        username = request.POST.get('UserName','')
        password = request.POST.get('Password','')
        user = auth.authenticate(username=username,password=password)
        if user is not None:
            auth.login(request,user)
            return redirect('/')
    context = {}
    return render(request, 'main/login.html', context)

def logout(request):
    auth.logout(request)
    return redirect('/')

def questionAnswer(request):
    questions = Question.objects.all()
    answers = Answer.objects.all()
    context = {
        'questions':questions,
        'answers':answers,
    }
    return render(request, 'main/question_answer.html',context)

def chooseExam(request):
    exams = Exam.objects.all()
    context = {
        'exams':exams,
        'firsExam':exams[0],
    }
    return render(request, 'main/choose_exam.html', context)

def testing(request):
    if('Exam' not in request.GET):
        return redirect('/choose_exam')
    thisExam = Exam.objects.get(id=request.GET['Exam'])
    allQuestions = Question.objects.all().filter(exam=thisExam)
    idies = []
    for quest in allQuestions:
        idies.append(quest.id)
    idies = choosing_questions.chosseQuestions(idies,thisExam.number_questions)
    arrQuest = Question.objects.all().filter(id__in=idies)
    arrAnswer = Answer.objects.all().filter(question__in=arrQuest)
    context = {
        'arrQuest':arrQuest,
        'arrAnswer':arrAnswer,
        'examId':thisExam.id,
    }
    return render(request,'main/testing.html',context)

def checkAnswer(request):
    numAnswers = []
    mark = 0
    for i in request.POST.keys():
        if i.find('quest') != -1:
            numAnswers.append(request.POST[i])
    answers = Answer.objects.all().filter(id__in=numAnswers)
    thisExam = Exam.objects.get(id=request.POST['examId'])
    for ans in answers:
        if ans.right:
            mark += 1
    test = Test.objects.all().filter(id=1)
    test = Test(user=request.user,exam=thisExam,time=datetime.datetime.now(),mark=mark)
    test.save()
    return redirect('/')

def showResults(request):
    dateFrom = None
    dateTo = None
    if request.method == 'GET' and ('dateFrom' in request.GET or 'dateTo' in request.GET):
        dateFrom = request.GET['dateFrom']
        dateTo = request.GET['dateTo']
        if dateTo == '':
            dateTo = '2200-01-01'
        if dateFrom == '':
            dateFrom = '1980-01-01'
        test = Test.objects.all().filter(time__range=(dateFrom, dateTo+" 23:59:59"))
    else:
        test = Test.objects.all()

    users = []
    for foo in test:
        users.append(foo.user)
    workers = Worker.objects.all().filter(user__in=users)

    if dateTo == '2100-01-01':
        dateTo = None
    if dateFrom == '1980-01-01':
        dateFrom = None

    context = {
        'test':test,
        'dateFrom':dateFrom,
        'dateTo':dateTo,
        'workers':workers,
    }
    return render(request, 'main/show_results.html', context)

def createFile(request):
    if request.method != 'POST' or 'dateTo' not in request.POST or 'dateFrom' not in request.POST:
       return redirect('/')
    dateTo = request.POST['dateTo']
    dateFrom = request.POST['dateFrom']
    if dateTo == 'None':
        dateTo = '2200-01-01'
    if dateFrom == 'None':
        dateFrom = '1980-01-01'
    test = Test.objects.all().filter(time__range=(dateFrom, dateTo+" 23:59:59"))

    users = []
    for foo in test:
        users.append(foo.user)
    workers = Worker.objects.all().filter(user__in=users)

    #write data to xls file
    wb = openpyxl.Workbook()
    #wb.create_sheet('test')
    sheet = wb['Sheet']

    sheet['A1'] = '№'
    sheet['B1'] = 'Время'
    sheet['C1'] = 'Ползователь'
    sheet['D1'] = 'Имя'
    sheet['E1'] = 'Фамилия'
    sheet['F1'] = 'Номер'
    sheet['G1'] = 'Подразделение'
    sheet['H1'] = 'Экзамен'
    sheet['I1'] = 'Оценка'

    numRow = 2
    for foo in test:
        sheet['A'+str(numRow)] = str(foo.id)
        sheet['B'+str(numRow)] = foo.time.strftime('%Y-%m-%d %H:%M')
        sheet['C'+str(numRow)] = foo.user.username
        for worker in workers:
            if worker.user == foo.user:
                sheet['D' + str(numRow)] = worker.first_name
                sheet['E' + str(numRow)] = worker.last_name
                sheet['F' + str(numRow)] = worker.number
                sheet['G' + str(numRow)] = worker.subdivision.name

        sheet['H'+str(numRow)] = foo.exam.name
        sheet['I'+str(numRow)] = str(foo.mark)+'/'+str(foo.exam.number_questions)
        numRow += 1

    #sheet.Columns('A:E').EntireColumn.AutoFit()

    response = HttpResponse(content=save_virtual_workbook(wb))  #mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=result_test.xlsx'
    return response

def dawnloadQuestions(request):
    if request.method != 'POST' or 'QuestionFile' not in request.FILES:
        return redirect('/quest_ans')
    wb = openpyxl.load_workbook(request.FILES['QuestionFile'])
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    text = sheet['A1'].value
    return redirect('/quest_ans')
